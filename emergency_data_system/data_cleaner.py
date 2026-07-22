"""
================================================================================
  应急数据智能清洗引擎
================================================================================
  本模块是系统的核心业务逻辑层，基于 Pandas 与 openpyxl 实现完整的
  Excel 数据清洗流水线。

  架构设计:
    本模块与 Web 服务层（app.py）完全解耦，可三种方式运行：
      1. 作为 Flask 后端引擎：  from data_cleaner import clean_data
      2. 作为命令行工具：        python data_cleaner.py test.xlsx
      3. 作为 Python 库调用：    clean_data("/path/to/file.xlsx")

  清洗流水线（按执行顺序）:
    步骤1  读取Excel    → load_excel()          支持 .xlsx，自动剔除全空行列
    步骤2  列名标准化   → standardize_columns()  20+种字段变体 → 5类标准字段
    步骤3  日期规范化   → normalize_date()       6种格式 → YYYY-MM-DD
    步骤4  号码校验     → clean_phone()          手机/固话/国家码校验
    步骤5  名称规整     → clean_company_name()   去空格/统一括号/空值标记
    步骤6  地址补全     → complete_address()     地域前缀自动补全
    步骤7  去重+填充    → deduplicate_and_fill() 按维度去重+统一填充
    步骤8  异常日志     → collect_anomaly_logs() 4类异常检测+日志生成
    步骤9  保存结果     → 清洗后Excel + 异常日志Excel + 橙色字体标注

  异常类型（共4种，记录到异常日志中）:
    1. 必填字段缺失   —— 企业名称/联系电话/排查日期为空
    2. 联系方式格式错误 —— 手机号/固话格式不合法
    3. 企业名称缺失   —— 企业名称为空或无效占位符("无"/"暂无")
    4. 日期格式无法识别 —— 日期字符串无法解析，已填充为默认日期

  设计决策:
    - 所有值以字符串类型读入（dtype=str），避免 Pandas 隐式类型转换
      导致前导零丢失（如 010-88886666 → 1088886666）
    - 清洗规则全配置化（config.py），新增规则无需改代码
    - 辅助列机制（_phone_valid / _date_anomaly）追踪中间判断，
      输出前自动剥离，确保报表干净
================================================================================
"""

import re
import os
from datetime import datetime
from typing import Tuple, List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from config import (
    COLUMN_STANDARD_MAP,
    REQUIRED_FIELDS,
    DEDUP_KEYS,
    FILL_VALUE,
    REGION_PREFIX,
    OUTPUT_FOLDER,
)


# ====================================================================
# 【小节3.1】Excel 文件读取
# ====================================================================

def load_excel(file_path: str) -> pd.DataFrame:
    """
    读取上传的 Excel 文件（.xlsx 格式）

    参数:
        file_path: Excel 文件路径

    返回:
        pd.DataFrame: 读取到的数据表

    异常:
        ValueError: 文件格式不支持或读取失败
    """
    if not os.path.exists(file_path):
        raise ValueError(f"文件不存在: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xlsx':
        df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx 格式")

    if df.empty:
        raise ValueError("文件内容为空，请检查上传文件")

    # 去除全空行和全空列
    df = df.dropna(how='all').dropna(axis=1, how='all')

    return df


# ====================================================================
# 【小节3.2】字段列名标准化
# ====================================================================

def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    智能统一差异化字段命名
    例：'企业'/'公司'/'单位' → '企业名称'
        '手机'/'电话'/'手机号' → '联系电话'

    参数:
        df: 原始 DataFrame

    返回:
        列名标准化后的 DataFrame
    """
    # 去除列名首尾空格
    df.columns = [str(col).strip() for col in df.columns]

    rename_map = {}
    for col in df.columns:
        if col in COLUMN_STANDARD_MAP:
            rename_map[col] = COLUMN_STANDARD_MAP[col]

    if rename_map:
        df = df.rename(columns=rename_map)

    return df


# ====================================================================
# 【小节3.3】日期格式规范化
# ====================================================================

def normalize_date(value: Any, default_date: str = None) -> str:
    """
    将各种不规则日期格式统一为 YYYY-MM-DD

    支持的输入格式示例：
        - 2024/1/5
        - 2024年1月5日
        - 2024.1.5
        - 20240105
        - 1/5/2024
        - 空值/NaN → 用默认日期填充

    参数:
        value: 原始日期值
        default_date: 默认日期（如上报当日），格式 YYYY-MM-DD

    返回:
        标准化后的日期字符串 YYYY-MM-DD
    """
    # 默认日期：如果调用方未指定，使用当天日期
    if default_date is None:
        default_date = datetime.now().strftime('%Y-%m-%d')

    # 空值检测：NaN / None / 空字符串 / 字符串'nan' → 返回默认日期
    if pd.isna(value) or str(value).strip() == '' or str(value).strip() in ('nan', 'NaN', 'None', ''):
        return default_date

    date_str = str(value).strip()

    # ---- 解析策略：逐级匹配，命中即返回 ----
    # 设计理由：不依赖 pd.to_datetime() 的隐式推断（其行为在不同 Pandas
    #          版本间不一致），而是使用显式的正则匹配，结果可预测。

    # 策略1: 中文格式 → "2024年1月5日" 或 "2024年01月05日"
    # 正则说明：4位年份 + "年" + 1-2位月份 + "月" + 1-2位日期 + 可选的"日"
    chinese_match = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日?', date_str)
    if chinese_match:
        y, m, d = chinese_match.groups()
        return f'{int(y):04d}-{int(m):02d}-{int(d):02d}'

    # 统一分隔符：将 "." 和 "/" 替换为 "-"，便于后续统一处理
    date_str_clean = re.sub(r'[./]', '-', date_str)

    # 策略2: 纯8位数字 → "20240105"（基层模板中常见的无分隔符格式）
    if date_str_clean.isdigit() and len(date_str_clean) == 8:
        y, m, d = date_str_clean[:4], date_str_clean[4:6], date_str_clean[6:8]
        return f'{int(y):04d}-{int(m):02d}-{int(d):02d}'

    # 策略3: 标准分隔格式 → "2024-1-5" 或 "1-5-2024"（美式）
    # 依次尝试两种正则模式，命中后验证日期合法性
    patterns = [
        # (正则, 模式)  mode=None → 年-月-日; mode='md' → 月-日-年
        (r'^(\d{4})-(\d{1,2})-(\d{1,2})$', None),   # 2024-1-5
        (r'^(\d{1,2})-(\d{1,2})-(\d{4})$', 'md'),    # 1-5-2024（美式）
    ]

    for pattern, mode in patterns:
        match = re.match(pattern, date_str_clean)
        if match:
            g = match.groups()
            if mode == 'md':
                m, d, y = g  # 美式：月/日/年
            else:
                y, m, d = g  # 标准：年/月/日
            try:
                y, m, d = int(y), int(m), int(d)
                # 合理范围校验：年2000-2100，月1-12，日1-31
                if 2000 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
                    # 进一步验证日期合法性（排除 2月30日 等非法日期）
                    # datetime() 对非法日期会抛出 ValueError
                    datetime(y, m, d)
                    return f'{y:04d}-{m:02d}-{d:02d}'
            except (ValueError, OverflowError):
                # 日期不合法（如 2024-02-30），继续尝试下一个模式
                pass

    # 所有解析策略均失败 → 返回默认日期
    # 注意：此情况会被 normalize_dates_in_df 中的 _date_anomaly 标记捕获
    return default_date


def normalize_dates_in_df(df: pd.DataFrame, date_column: str = '排查日期') -> pd.DataFrame:
    """
    对 DataFrame 中指定日期列进行批量规范化，并标记日期异常

    参数:
        df: 数据表
        date_column: 日期列名

    返回:
        日期列已规范化的 DataFrame（附加 _date_anomaly 和 _date_original 辅助列）
    """
    if date_column not in df.columns:
        df[date_column] = datetime.now().strftime('%Y-%m-%d')
        return df

    # 保存原始值用于异常检测
    original_values = df[date_column].copy()
    default = datetime.now().strftime('%Y-%m-%d')
    df[date_column] = df[date_column].apply(lambda v: normalize_date(v, default))

    # 标记日期异常：原始值非空但无法解析的（被替换为默认日期的）
    df['_date_anomaly'] = False
    df['_date_original'] = ''
    for i in df.index:
        orig = str(original_values[i]).strip()
        if orig and orig not in ('nan', 'NaN', 'None', ''):
            # 检查 normalize_date 是否成功解析
            test_result = normalize_date(orig, '__TEST__')
            if test_result == '__TEST__':
                df.at[i, '_date_anomaly'] = True
                df.at[i, '_date_original'] = orig

    return df


# ====================================================================
# 【小节3.4】联系方式智能校验
# ====================================================================

def clean_phone(phone: Any) -> Tuple[str, bool, str]:
    """
    清洗并校验电话号码

    处理流程：
        1. 去除空格、横线、括号等无效字符
        2. 校验是否为合法手机号（11位，1开头）
        3. 校验是否为合法固话（区号+号码）

    参数:
        phone: 原始电话号码

    返回:
        (清洗后号码, 是否合法, 异常原因)
    """
    if pd.isna(phone) or str(phone).strip() == '':
        return (FILL_VALUE, False, '联系电话为空')

    raw = str(phone).strip()

    # 清除无效字符：空格、横线（中英文）、括号（中英文）、点号
    # 保留数字和 '+'（用于国际号码前缀 +86）
    cleaned = re.sub(r'[\s\-\(\)（）\-\.]', '', raw)

    # 清洗后为空 → 原始值全是无效字符
    if not cleaned:
        return (FILL_VALUE, False, '联系电话为空')

    # 校验策略（按优先级依次尝试）:

    # 策略1: 中国大陆手机号 —— 11位，以1开头，第二位为3-9
    if re.match(r'^1[3-9]\d{9}$', cleaned):
        return (cleaned, True, '')

    # 策略2: 固话（含区号） —— 0开头 + 3-4位区号 + 7-8位号码
    # 示例: 01088886666 (北京10位), 075588886666 (深圳11位)
    if re.match(r'^0\d{2,3}\d{7,8}$', cleaned):
        return (cleaned, True, '')

    # 策略3: 带国家码的手机号 —— +86 开头
    # 去除 +86 前缀，保留后11位作为标准手机号
    if re.match(r'^\+86\d{11}$', cleaned):
        return (cleaned[3:], True, '')

    # 所有策略失败 → 标记为异常，保留原始值供人工复核
    return (raw, False, f'联系电话格式不正确: {raw}')


def clean_phones_in_df(df: pd.DataFrame, phone_column: str = '联系电话') -> pd.DataFrame:
    """
    批量清洗 DataFrame 中的联系电话列

    参数:
        df: 数据表
        phone_column: 电话列名

    返回:
        电话列已清洗的 DataFrame，异常信息记录在 _phone_valid 和 _phone_error 辅助列
    """
    if phone_column not in df.columns:
        return df

    results = df[phone_column].apply(clean_phone)
    df[phone_column] = results.apply(lambda r: r[0])
    df['_phone_valid'] = results.apply(lambda r: r[1])
    df['_phone_error'] = results.apply(lambda r: r[2])

    return df


# ====================================================================
# 【小节3.5】企业名称规整
# ====================================================================

def clean_company_name(name: Any) -> Tuple[str, bool, str]:
    """
    规整企业名称：去空格、统一括号、检查空值

    参数:
        name: 原始企业名称

    返回:
        (规整后名称, 是否有效, 异常原因)
    """
    if pd.isna(name) or str(name).strip() == '':
        return (FILL_VALUE, False, '企业名称为空')

    cleaned = str(name).strip()

    # 统一中英文括号：全角 → 半角（保持一致性）
    # 注：中文环境通常保留全角括号，但统一为全角更规范
    cleaned = cleaned.replace('(', '（').replace(')', '）')

    # 去除多余空格
    cleaned = re.sub(r'\s+', '', cleaned)

    if not cleaned or cleaned in ('nan', 'NaN', 'None', '无', '暂无'):
        return (FILL_VALUE, False, f'企业名称为无效值: {name}')

    return (cleaned, True, '')


# ====================================================================
# 【小节3.6】地址自适应补全
# ====================================================================

def complete_address(address: Any) -> str:
    """
    自动补全缺失地域前缀的地址

    根据 config.py 中的 REGION_PREFIX 配置，
    自动为缺少省/市/区前缀的地址补全地域信息

    参数:
        address: 原始地址字符串

    返回:
        补全后的地址
    """
    if pd.isna(address) or str(address).strip() == '':
        return FILL_VALUE

    addr = str(address).strip()

    # 拼接配置中的地域前缀
    prefix_parts = []
    for key in ['province', 'city', 'district']:
        val = REGION_PREFIX.get(key, '')
        if val:
            prefix_parts.append(val)

    full_prefix = ''.join(prefix_parts)

    if not full_prefix:
        return addr  # 未配置前缀，原样返回

    # 如果地址已经包含前缀，不重复添加
    if addr.startswith(full_prefix):
        return addr

    # 逐级检查
    for key in ['district', 'city', 'province']:
        val = REGION_PREFIX.get(key, '')
        if val and not addr.startswith(val):
            # 找到第一个缺失的级别，补全它及后面的所有级别
            prefix_parts = []
            for k in ['province', 'city', 'district']:
                v = REGION_PREFIX.get(k, '')
                if v and (k == key or not addr.startswith(v)):
                    prefix_parts.append(v)
            full_prefix = ''.join(prefix_parts)
            return full_prefix + addr

    return addr


def complete_addresses_in_df(df: pd.DataFrame, addr_column: str = '企业地址') -> pd.DataFrame:
    """
    批量补全地址列

    参数:
        df: 数据表
        addr_column: 地址列名

    返回:
        地址已补全的 DataFrame
    """
    if addr_column in df.columns:
        df[addr_column] = df[addr_column].apply(complete_address)
    return df


# ====================================================================
# 【小节3.6】数据去重 + 缺失填充
# ====================================================================

def deduplicate_and_fill(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """
    以核心维度去重，并对空白字段统一填充

    去重维度：企业名称 + 排查日期（可在 config.py 的 DEDUP_KEYS 中修改）
    填充值：'待补充'（可在 config.py 的 FILL_VALUE 中修改）

    参数:
        df: 数据表

    返回:
        (去重填充后的 DataFrame, 被剔除的重复行数)
    """
    original_count = len(df)

    # 去重：保留第一次出现的记录
    dedup_keys = [k for k in DEDUP_KEYS if k in df.columns]
    if dedup_keys:
        df = df.drop_duplicates(subset=dedup_keys, keep='first')

    removed_count = original_count - len(df)

    # 空白字段统一填充
    df = df.fillna(FILL_VALUE)
    # 空字符串也填充
    df = df.replace(r'^\s*$', FILL_VALUE, regex=True)

    return df, removed_count


# ====================================================================
# 【小节3.6】异常日志收集
# ====================================================================

def collect_anomaly_logs(df: pd.DataFrame) -> List[Dict[str, str]]:
    """
    收集所有异常数据记录，生成清洗日志

    检查项：
        1. 必填字段为空/填充值
        2. 电话号码不合法
        3. 企业名称为填充值
        4. 日期格式异常（已由 normalize_date 处理，此处检测残留）

    参数:
        df: 清洗后的 DataFrame

    返回:
        异常日志列表，每条格式: {行号, 字段, 原始值, 异常类型, 处理方式}
    """
    logs = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel 行号（+2 补偿0-based和表头）

        # 检查必填字段
        for field in REQUIRED_FIELDS:
            if field in df.columns:
                val = row.get(field, '')
                if str(val).strip() == FILL_VALUE or str(val).strip() == '':
                    logs.append({
                        '行号': row_num,
                        '字段': field,
                        '原始值': str(val),
                        '异常类型': '必填字段缺失',
                        '处理方式': f'已填充为"{FILL_VALUE}"'
                    })

        # 检查电话号码（来源于 _phone_valid 辅助列）
        if '_phone_valid' in df.columns:
            is_valid = row.get('_phone_valid', True)
            if str(is_valid) == 'False':
                logs.append({
                    '行号': row_num,
                    '字段': '联系电话',
                    '原始值': row.get('_phone_error', ''),
                    '异常类型': '联系方式格式错误',
                    '处理方式': '已标记，建议人工复核修改'
                })

        # 检查企业名称
        if '企业名称' in df.columns:
            name = row.get('企业名称', '')
            if str(name).strip() == FILL_VALUE:
                logs.append({
                    '行号': row_num,
                    '字段': '企业名称',
                    '原始值': '(空)',
                    '异常类型': '企业名称缺失',
                    '处理方式': f'已填充为"{FILL_VALUE}"，建议人工补录'
                })

        # 检查日期异常
        if '_date_anomaly' in df.columns:
            is_date_anomaly = row.get('_date_anomaly', False)
            if str(is_date_anomaly) == 'True':
                logs.append({
                    '行号': row_num,
                    '字段': '排查日期',
                    '原始值': row.get('_date_original', ''),
                    '异常类型': '日期格式无法识别',
                    '处理方式': '已填充为上报当日日期，建议人工复核'
                })

    return logs


def save_anomaly_logs(logs: List[Dict], file_path: str = None) -> str:
    """
    将异常日志保存为 Excel 文件

    参数:
        logs: 异常日志列表
        file_path: 保存路径（可选，默认自动生成带时间戳的文件名）

    返回:
        日志文件路径
    """
    if file_path is None:
        from config import LOG_FOLDER
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_path = os.path.join(LOG_FOLDER, f'清洗日志_{timestamp}.xlsx')

    if logs:
        log_df = pd.DataFrame(logs)
        log_df.to_excel(file_path, index=False, engine='openpyxl')
    else:
        # 无异常也生成日志（空日志）
        log_df = pd.DataFrame([{'备注': '本次清洗未发现异常数据'}])
        log_df.to_excel(file_path, index=False, engine='openpyxl')

    return file_path


# ====================================================================
# 【整合入口】一键清洗：调用上述所有子模块
# ====================================================================

def clean_data(file_path: str) -> Dict[str, Any]:
    """
    一键智能清洗入口函数 —— 系统核心流水线

    本函数编排了完整的 9 步清洗流水线，每一步独立封装、顺序调用。
    流水线设计允许在任意步骤之间插入新的处理逻辑，而无需重构。

    完整执行流程（9步）:
        ┌──────────────────────────────────────────────────┐
        │ 步骤1  读取 Excel     → df                        │
        │ 步骤2  列名标准化      → df（统一5类字段命名）        │
        │ 步骤3  日期规范化      → df（6种格式→YYYY-MM-DD）    │
        │ 步骤4  号码校验       → df（+_phone_valid辅助列）   │
        │ 步骤5  名称规整       → df（去空格+统一括号+空值标记）│
        │ 步骤6  地址补全       → df（地域前缀自动拼接）        │
        │ 步骤7  去重 + 填充    → df（按维度去重+空白统一填充）  │
        │ 步骤8  异常日志收集   → logs（4类异常检测）          │
        │ 步骤9  保存结果       → Excel + 异常日志 + 橙色标注  │
        └──────────────────────────────────────────────────┘

    辅助列机制:
        步骤4 在 DataFrame 中增加 _phone_valid（布尔）和 _phone_error（字符串）
        步骤3 在 DataFrame 中增加 _date_anomaly（布尔）和 _date_original（字符串）
        这些辅助列在步骤9保存前统一剔除，确保输出报表干净无污染

    错误处理:
        - 文件不存在 / 空文件 / 格式不支持 → 抛出 ValueError
        - 清洗过程中某步失败 → 异常向上传播，由调用方（app.py）捕获并返回500
        - 日期解析失败 → 静默降级为默认日期，通过辅助列标记，不阻断流水线

    参数:
        file_path: str —— 原始 Excel 文件的绝对或相对路径

    返回:
        Dict[str, Any]:
            original_count      int    原始数据总条数
            valid_count         int    去重后的有效数据条数
            removed_dup_count   int    被剔除的重复数据条数
            anomaly_count       int    包含异常的数据行数（去重计数）
            output_path         str    清洗后 Excel 文件的绝对路径
            log_path            str    异常日志 Excel 文件的绝对路径
            columns             list   清洗后的标准列名列表
            anomaly_logs        list   异常日志明细（用于前端展示和数据库存储）
    """
    # ================================================================
    #  步骤1: 读取 Excel 文件
    # ================================================================
    #  load_excel 以 dtype=str 读入所有数据，防止电话号码前导零丢失
    #  全空行和全空列自动剔除
    df = load_excel(file_path)
    original_count = len(df)  # 记录原始条数（用于统计和有效率计算）

    # ================================================================
    #  步骤2: 字段列名标准化
    # ================================================================
    #  将各基层单位上报的不同列名（企业/公司/手机/电话...）统一映射为
    #  5 个标准字段名：企业名称、联系电话、排查日期、企业地址、排查类型
    df = standardize_columns(df)

    # ================================================================
    #  步骤3: 日期格式规范化
    # ================================================================
    #  寻找日期列：优先匹配"排查日期"，其次匹配含"日期"的任意列名
    #  如果都没有，跳过日期处理（如完全不相关的Excel文件）
    if '排查日期' in df.columns or any('日期' in c for c in df.columns):
        date_col = '排查日期' if '排查日期' in df.columns else None
        if date_col is None:
            for c in df.columns:
                if '日期' in c:
                    date_col = c
                    break
        # normalize_dates_in_df 返回带 _date_anomaly 和 _date_original 的 DataFrame
        df = normalize_dates_in_df(df, date_col or '排查日期')

    # ================================================================
    #  步骤4: 联系电话智能校验
    # ================================================================
    #  仅当存在"联系电话"列时执行；clean_phones_in_df 返回带 _phone_valid
    #  和 _phone_error 辅助列的 DataFrame
    if '联系电话' in df.columns:
        df = clean_phones_in_df(df, '联系电话')

    # ================================================================
    #  步骤5: 企业名称规整
    # ================================================================
    #  优先处理已命名为"企业名称"的列
    #  如果不存在，智能搜索含"名称""企业""公司""单位"的列并重命名
    if '企业名称' in df.columns:
        name_results = df['企业名称'].apply(clean_company_name)
        df['企业名称'] = name_results.apply(lambda r: r[0])
    else:
        # 智能匹配：在列名中搜索关键词
        name_col = None
        for c in df.columns:
            if '名称' in c or '企业' in c or '公司' in c or '单位' in c:
                name_col = c
                break
        if name_col:
            df = df.rename(columns={name_col: '企业名称'})
            df['企业名称'] = df['企业名称'].apply(lambda n: clean_company_name(n)[0])

    # ================================================================
    #  步骤6: 地址自适应补全
    # ================================================================
    #  仅当存在"企业地址"列且 REGION_PREFIX 配置了前缀时生效
    if '企业地址' in df.columns:
        df = complete_addresses_in_df(df, '企业地址')

    # ================================================================
    #  步骤7: 数据去重 + 缺失值填充
    # ================================================================
    #  去重维度：企业名称 + 排查日期（在 config.py 的 DEDUP_KEYS 中配置）
    #  填充策略：NaN / None / '' → FILL_VALUE（默认"待补充"）
    #  removed_count 记录被剔除的重复行数，用于统计展示
    df, removed_count = deduplicate_and_fill(df)

    # ================================================================
    #  步骤8: 异常日志收集
    # ================================================================
    #  检测 4 类异常：必填字段缺失 / 联系方式格式错误 / 企业名称缺失 / 日期无法识别
    #  每条异常包含：行号、字段名、原始值、异常类型、处理方式
    logs = collect_anomaly_logs(df)

    # 计算异常条数（按行号去重：同一行有多个异常字段时只计 1 条）
    anomaly_rows = set(log['行号'] for log in logs)
    anomaly_count = len(anomaly_rows)

    # ---- 保存清洗结果 ----
    # 移除辅助列
    output_df = df.drop(columns=['_phone_valid', '_phone_error', '_date_anomaly', '_date_original'], errors='ignore')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'清洗后数据_{timestamp}.xlsx'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    output_df.to_excel(output_path, index=False, engine='openpyxl')

    # ---- 标注"待补充"单元格为黄色字体 ----
    YELLOW_FONT = Font(color='FFA500', bold=True)
    wb = load_workbook(output_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and str(cell.value).strip() == FILL_VALUE:
                cell.font = YELLOW_FONT
    wb.save(output_path)

    # ---- 保存异常日志 ----
    log_path = save_anomaly_logs(logs)

    # ---- 返回统计信息 ----
    return {
        'original_count': original_count,
        'valid_count': len(output_df),
        'removed_dup_count': removed_count,
        'anomaly_count': anomaly_count,
        'output_path': output_path,
        'log_path': log_path,
        'columns': list(output_df.columns),
        'anomaly_logs': logs,
    }


# ====================================================================
# 【独立测试入口】直接运行此文件可测试清洗引擎
# ====================================================================

if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print('用法: python data_cleaner.py <Excel文件路径>')
        print('示例: python data_cleaner.py test_data.xlsx')
        sys.exit(1)

    test_file = sys.argv[1]
    print(f'正在清洗: {test_file}')
    print('=' * 50)

    result = clean_data(test_file)

    print(f'原始数据条数: {result["original_count"]}')
    print(f'有效数据条数: {result["valid_count"]}')
    print(f'剔除重复条数: {result["removed_dup_count"]}')
    print(f'异常数据条数: {result["anomaly_count"]}')
    print(f'清洗后文件:   {result["output_path"]}')
    print(f'异常日志文件: {result["log_path"]}')
    print(f'数据列名:     {result["columns"]}')
    print('=' * 50)
    print('✅ 清洗完成！')
