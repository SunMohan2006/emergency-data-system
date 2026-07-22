"""
================================================================================
  增强报表导出模块
================================================================================
  基于 openpyxl 生成多 Sheet 格式化合规报表，每个报表包含三个 Sheet：
     Sheet 1「清洗数据」  —— 清洗后的完整数据表（带格式和条件标注）
     Sheet 2「异常日志」  —— 异常数据明细（带黄色高亮标注）
     Sheet 3「统计汇总」  —— 数据质量概览与有效率计算

  格式化特性:
    1. 表头样式
       —— 深蓝色背景（#1A3A5C）+ 白色加粗字体（微软雅黑11pt）
       —— 表头行自动冻结（冻结窗格），滚动数据时表头保持可见

    2. 条件标注
       —— "待补充"单元格：橙色加粗字体（#FFA500）+ 浅黄色背景（#FFF3CD）
       —— 有效数据条数：绿色大字
       —— 异常数据条数：红色大字
       —— 异常日志中涉及"待补充"的行同样标注

    3. 列宽自适应
       —— 中文字符按 2 个字符宽度计算，英文按 1 个字符
       —— 宽度范围限制在 10~40 字符之间，防止极端值

    4. 数据质量率
       —— 统计汇总 Sheet 中展示"数据有效率 = 有效条数 / 原始条数"
       —— 直观反映该批次数据的整体质量水平

  设计决策:
    - 纯 openpyxl 实现（无 pandas.ExcelWriter），获得更精细的样式控制
    - 所有样式定义为模块级常量，便于统一调整品牌色
    - 增强报表生成失败不影响主流程（app.py 中以 try-except 包裹）

  使用方式:
    from report import export_full_report, export_clean_data_only

    # 完整报表（3 Sheet）
    export_full_report(data_df, anomaly_logs, stats, output_path, filename)

    # 仅数据 Sheet（向后兼容原有的简单下载）
    export_clean_data_only(data_df, output_path)
================================================================================
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==================== 样式常量 ====================

# 颜色定义
HEADER_FILL = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1A3A5C')
SUBTITLE_FONT = Font(name='微软雅黑', size=10, color='666666')
DATA_FONT = Font(name='微软雅黑', size=10)
ANOMALY_FONT = Font(name='微软雅黑', size=10, color='FFA500', bold=True)
ANOMALY_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

# 边框
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

# 对齐
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 填充标记值
FILL_MARKER = '待补充'


# ==================== 核心导出函数 ====================

def export_full_report(
    data_df: pd.DataFrame,
    anomaly_logs: List[Dict[str, str]],
    stats: Dict[str, Any],
    output_path: str,
    source_filename: str = '',
) -> str:
    """
    导出完整报表（多 Sheet）

    参数:
        data_df: 清洗后的数据 DataFrame
        anomaly_logs: 异常日志列表
        stats: 统计信息字典（来自 clean_data 返回值）
        output_path: 输出文件路径（.xlsx）
        source_filename: 原始上传文件名（用于报表说明）

    返回:
        输出文件路径

    Sheet 结构:
        Sheet 1「清洗数据」—— 清洗后的完整数据表
        Sheet 2「异常日志」—— 异常数据明细
        Sheet 3「统计汇总」—— 数据质量概览
    """
    wb = Workbook()

    # ---- Sheet 1: 清洗数据 ----
    ws_data = wb.active
    ws_data.title = '清洗数据'
    _write_data_sheet(ws_data, data_df, source_filename)

    # ---- Sheet 2: 异常日志 ----
    ws_log = wb.create_sheet('异常日志')
    _write_log_sheet(ws_log, anomaly_logs, source_filename)

    # ---- Sheet 3: 统计汇总 ----
    ws_stats = wb.create_sheet('统计汇总')
    _write_stats_sheet(ws_stats, stats, source_filename)

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return output_path


def _write_data_sheet(ws, df: pd.DataFrame, source: str) -> None:
    """写入清洗数据 Sheet"""
    # 标题行
    ws.merge_cells('A1:F1')
    ws['A1'] = '应急数据清洗结果'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN

    # 副标题
    ws.merge_cells('A2:F2')
    ws['A2'] = f'原始文件: {source or "未知"}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = CENTER_ALIGN

    # 空行
    ws.append([])

    # 表头（从第4行开始）
    header_row = 4
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            # 处理 NaN / None
            if pd.isna(val):
                val = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

            # "待补充"单元格橙色加粗
            if str(val).strip() == FILL_MARKER:
                cell.font = ANOMALY_FONT
                cell.fill = ANOMALY_FILL

    # 冻结首行（表头）
    ws.freeze_panes = f'A{header_row + 1}'

    # 自适应列宽
    _auto_fit_columns(ws, len(df.columns))


def _write_log_sheet(ws, logs: List[Dict[str, str]], source: str) -> None:
    """写入异常日志 Sheet"""
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = '数据清洗异常日志'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN

    ws.merge_cells('A2:E2')
    ws['A2'] = f'原始文件: {source or "未知"}  |  异常总数: {len(logs)} 条'
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = CENTER_ALIGN

    ws.append([])

    # 表头
    log_headers = ['Excel行号', '异常字段', '原始值', '异常类型', '处理方式']
    header_row = 4
    for col_idx, header in enumerate(log_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    if logs:
        keys_map = {'行号': 'Excel行号', '字段': '异常字段', '原始值': '原始值',
                     '异常类型': '异常类型', '处理方式': '处理方式'}
        for row_idx, log in enumerate(logs, start=header_row + 1):
            for col_idx, (log_key, _) in enumerate(keys_map.items(), start=1):
                val = log.get(log_key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER

                if str(val).strip() == FILL_MARKER or '待补充' in str(val):
                    cell.font = ANOMALY_FONT
                    cell.fill = ANOMALY_FILL
    else:
        ws.merge_cells(f'A{header_row + 1}:E{header_row + 1}')
        ws.cell(row=header_row + 1, column=1,
                value='本次清洗未发现异常数据').font = Font(name='微软雅黑', size=10, color='27AE60')

    # 冻结 + 列宽
    ws.freeze_panes = f'A{header_row + 1}'
    _auto_fit_columns(ws, 5)


def _write_stats_sheet(ws, stats: Dict[str, Any], source: str) -> None:
    """写入统计汇总 Sheet"""
    # 标题
    ws.merge_cells('A1:B1')
    ws['A1'] = '数据质量统计汇总'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN

    ws.merge_cells('A2:B2')
    ws['A2'] = f'原始文件: {source or "未知"}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = CENTER_ALIGN

    ws.append([])

    # 统计指标
    indicators = [
        ('原始数据条数', stats.get('original_count', 0)),
        ('有效数据条数', stats.get('valid_count', 0)),
        ('剔除重复条数', stats.get('removed_dup_count', 0)),
        ('异常数据条数', stats.get('anomaly_count', 0)),
    ]

    # 表头
    header_row = 4
    for col_idx, header in enumerate(['指标', '数值'], start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 指标数据
    for row_offset, (label, value) in enumerate(indicators, start=1):
        row = header_row + row_offset

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name='微软雅黑', size=10, bold=True)
        label_cell.alignment = LEFT_ALIGN
        label_cell.border = THIN_BORDER

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = Font(name='微软雅黑', size=14, bold=True)

        # 颜色区分
        if label == '有效数据条数':
            value_cell.font = Font(name='微软雅黑', size=14, bold=True, color='27AE60')
        elif label == '异常数据条数':
            value_cell.font = Font(name='微软雅黑', size=14, bold=True, color='E74C3C')
        elif label == '剔除重复条数':
            value_cell.font = Font(name='微软雅黑', size=14, bold=True, color='F39C12')

        value_cell.alignment = CENTER_ALIGN
        value_cell.border = THIN_BORDER

    # 数据质量率
    summary_row = header_row + len(indicators) + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    original = stats.get('original_count', 1) or 1
    valid = stats.get('valid_count', 0)
    quality_rate = valid / original * 100 if original > 0 else 0
    ws.cell(row=summary_row, column=1,
            value=f'数据有效率: {quality_rate:.1f}% ({valid}/{original})')
    ws.cell(row=summary_row, column=1).font = Font(name='微软雅黑', size=12, bold=True, color='1A3A5C')
    ws.cell(row=summary_row, column=1).alignment = CENTER_ALIGN

    # 列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def _auto_fit_columns(ws, max_col: int, min_width: int = 10, max_width: int = 40) -> None:
    """自动调整列宽（基于内容长度）"""
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val:
                    # 中文字符按2个宽度计算
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell_val))
                    max_length = max(max_length, length)
        adjusted = max(min_width, min(max_length + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted


# ==================== 快捷导出函数 ====================

def export_clean_data_only(df: pd.DataFrame, output_path: str) -> str:
    """
    仅导出清洗数据（兼容原有下载功能）

    参数:
        df: 清洗后的 DataFrame
        output_path: 输出路径

    返回:
        输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '清洗数据'

    # 写表头
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 写数据
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

            if str(val).strip() == FILL_MARKER:
                cell.font = ANOMALY_FONT
                cell.fill = ANOMALY_FILL

    ws.freeze_panes = 'A2'
    _auto_fit_columns(ws, len(df.columns))

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ==================== 独立测试入口 ====================

if __name__ == '__main__':
    import pandas as pd

    print('=' * 50)
    print('  报表导出模块测试')
    print('=' * 50)

    # 构造测试数据
    test_df = pd.DataFrame({
        '企业名称': ['华为技术', '待补充', '阿里巴巴'],
        '联系电话': ['13800138000', '12345', '待补充'],
        '排查日期': ['2024-01-05', '2024-02-15', '2024-03-20'],
        '企业地址': ['深圳', '待补充', '上海浦东'],
        '排查类型': ['安全生产', '待补充', '消防检查'],
    })

    test_logs = [
        {'行号': '3', '字段': '企业名称', '原始值': '(空)', '异常类型': '企业名称缺失', '处理方式': '已填充为"待补充"'},
        {'行号': '3', '字段': '联系电话', '原始值': '12345', '异常类型': '联系方式格式错误', '处理方式': '已标记'},
        {'行号': '4', '字段': '联系电话', '原始值': '(空)', '异常类型': '必填字段缺失', '处理方式': '已填充为"待补充"'},
    ]

    test_stats = {
        'original_count': 3,
        'valid_count': 3,
        'removed_dup_count': 0,
        'anomaly_count': 2,
    }

    # 生成报表
    output_path = 'outputs/_test_report.xlsx'
    export_full_report(test_df, test_logs, test_stats, output_path, 'test.xlsx')

    # 验证
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet_names = wb.sheetnames
    print(f'Sheet 数量: {len(sheet_names)} - {sheet_names}')

    ws_data = wb['清洗数据']
    print(f'数据Sheet: {ws_data.max_row} 行 x {ws_data.max_column} 列')

    # 检查黄色标注
    yellow_count = 0
    for row in ws_data.iter_rows(min_row=5):
        for cell in row:
            if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
                try:
                    if cell.font.color.rgb == '00FFA500':
                        yellow_count += 1
                except TypeError:
                    pass
    print(f'橙色标注单元格: {yellow_count} 个')

    # 清理
    os.remove(output_path)
    print('测试文件已清理')
    print()
    print('=' * 50)
    print('报表模块测试通过')
